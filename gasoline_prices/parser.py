from datetime import datetime
from io import BytesIO
from typing import BinaryIO

from requests import get as requests_get
from openpyxl import load_workbook

HI_OCTANE_ROWS: range = range(10, 16)
REGULAR_ROWS: range = range(16, 22)
DATE_COL: int = 4  # "D"
PRICE_COL: int = 15  # "O"
REF_COL: int = 16  # "P"


class Parser:
    def __init__(self, excel_url: str, user_agent: str) -> None:
        self.excel_url: str = excel_url
        self.user_agent: str = user_agent

    def fetch_excel_file(self) -> bool:
        headers = {"User-Agent": self.user_agent}
        print(f"GET {self.excel_url}")
        excel_data: bytes = requests_get(self.excel_url, headers=headers).content
        self.excel_data: BinaryIO = BytesIO(excel_data)

        return True

    def parse_excel_file(self) -> dict:
        prices: dict = dict()
        prices["hi_octane"] = self.__create_prices_by_type("hi_octane", HI_OCTANE_ROWS)
        prices["regular"] = self.__create_prices_by_type("regular", REGULAR_ROWS)

        self.excel_data.close()
        return prices

    def __create_prices_by_type(self, oil_type, target_rows) -> list:
        self.excel_data.seek(0)
        wb = load_workbook(self.excel_data)
        ws = wb.worksheets[0]

        prices_by_type: list = list()
        for row in target_rows:
            survey_date = ws.cell(row=row, column=DATE_COL).value.date()
            price = ws.cell(row=row, column=PRICE_COL).value
            ref_price = ws.cell(row=row, column=REF_COL).value
            prices_by_type.append({"survey_date": survey_date, "price": price, "ref_price": ref_price})

        return prices_by_type
